"""Excel export routes (openpyxl)."""
from fastapi import APIRouter, Depends, Response, Query
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from auth import require_roles, get_current_user
from db import db
from datetime import datetime, timezone

router = APIRouter(prefix="/exports", tags=["exports"])

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="E63946", end_color="E63946", fill_type="solid")
CENTER = Alignment(horizontal="center", vertical="center")


def _apply_header(ws, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
    ws.freeze_panes = "A2"


def _autosize(ws):
    for col in ws.columns:
        max_len = 8
        letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_len + 2, 60)


def _wb_response(wb: Workbook, filename: str) -> Response:
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/tasks.xlsx")
async def export_tasks(user=Depends(require_roles("super_admin", "admin", "manager"))):
    tasks = await db.tasks.find({}, {"_id": 0}).sort("created_at", -1).to_list(5000)
    users = {u["id"]: u for u in await db.users.find({}, {"_id": 0}).to_list(500)}
    projects = {p["id"]: p for p in await db.projects.find({}, {"_id": 0}).to_list(500)}
    sess_by_task = {}
    for s in await db.timer_sessions.find({}, {"_id": 0}).to_list(50000):
        sess_by_task[s["task_id"]] = sess_by_task.get(s["task_id"], 0) + (s.get("duration_seconds", 0) or 0)
    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"
    headers = ["Task ID", "Title", "Project", "Company", "Assignee", "Designation",
               "Priority", "Status", "Scheduled", "Due", "Estimate (min)",
               "Time spent (h)", "Created"]
    _apply_header(ws, headers)
    for t in tasks:
        a = users.get(t.get("assignee_id"))
        p = projects.get(t.get("project_id"))
        secs = sess_by_task.get(t["id"], 0)
        ws.append([
            t["id"], t.get("title", ""), (p or {}).get("name", ""),
            (p or {}).get("company_name", ""),
            (a or {}).get("first_name", ""), (a or {}).get("designation", ""),
            t.get("priority", ""), t.get("status", ""),
            t.get("scheduled_start_date", ""), t.get("due_date", ""),
            t.get("estimated_duration_minutes", 0),
            round(secs / 3600, 2), t.get("created_at", ""),
        ])
    _autosize(ws)
    return _wb_response(wb, "raybotix-tasks.xlsx")


@router.get("/costs.xlsx")
async def export_costs(
    range: str = Query("month"),
    start: str = "",
    end: str = "",
    project_id: str = "",
    user_id: str = "",
    user=Depends(require_roles("super_admin", "admin")),
):
    """Excel export mirroring /analytics/costs (respects the same filters)."""
    from routes_analytics import _range_window, _employee_costs, _session_cost_and_seconds

    win_start, win_end, label = _range_window(range, start, end)
    cost_map = await _employee_costs()

    users = {u["id"]: u for u in await db.users.find({}, {"_id": 0}).to_list(1000)}
    tasks = {t["id"]: t for t in await db.tasks.find({}, {"_id": 0}).to_list(10000)}
    projects = {p["id"]: p for p in await db.projects.find({}, {"_id": 0}).to_list(1000)}
    sessions = await db.timer_sessions.find({
        "$or": [
            {"started_at": {"$gte": win_start.isoformat(), "$lte": win_end.isoformat()}},
            {"ended_at": None},
            {"ended_at": {"$gte": win_start.isoformat()}},
        ]
    }, {"_id": 0}).to_list(200000)

    def _pass_filters(sess, task):
        if user_id and sess.get("user_id") != user_id:
            return False
        if project_id and (not task or task.get("project_id") != project_id):
            return False
        return True

    wb = Workbook()
    # Sheet 1: Cost per Task
    ws = wb.active
    ws.title = "Cost per Task"
    _apply_header(ws, ["Range", "Task", "Project", "Assignee", "Designation",
                       "Hours", "Hourly (₹)", "Cost (₹)"])
    task_agg = {}
    for s in sessions:
        info = cost_map.get(s.get("user_id"))
        t = tasks.get(s.get("task_id"))
        if not info or not t:
            continue
        if not _pass_filters(s, t):
            continue
        cost, secs = _session_cost_and_seconds(s, info["hourly"], win_start, win_end)
        if secs <= 0:
            continue
        row = task_agg.setdefault(t["id"], {"cost": 0.0, "seconds": 0})
        row["cost"] += cost
        row["seconds"] += secs
    for tid, v in sorted(task_agg.items(), key=lambda kv: kv[1]["cost"], reverse=True):
        t = tasks.get(tid, {})
        a = users.get(t.get("assignee_id"))
        p = projects.get(t.get("project_id"))
        info = cost_map.get(t.get("assignee_id"))
        hourly = info["hourly"] if info else 0
        ws.append([
            label, t.get("title", ""), (p or {}).get("name", ""),
            (a or {}).get("first_name", ""), (a or {}).get("designation", ""),
            round(v["seconds"] / 3600.0, 2), round(hourly, 2), round(v["cost"], 2),
        ])
    _autosize(ws)

    # Sheet 2: Cost per Project
    ws2 = wb.create_sheet("Cost per Project")
    _apply_header(ws2, ["Project", "Company", "Hours", "Cost (₹)"])
    proj_agg = {}
    for s in sessions:
        info = cost_map.get(s.get("user_id"))
        t = tasks.get(s.get("task_id"))
        if not info or not t:
            continue
        if not _pass_filters(s, t):
            continue
        cost, secs = _session_cost_and_seconds(s, info["hourly"], win_start, win_end)
        if secs <= 0:
            continue
        pid = t.get("project_id") or "__none__"
        row = proj_agg.setdefault(pid, {"cost": 0.0, "seconds": 0})
        row["cost"] += cost
        row["seconds"] += secs
    for pid, v in sorted(proj_agg.items(), key=lambda kv: kv[1]["cost"], reverse=True):
        p = projects.get(pid, {})
        ws2.append([
            p.get("name", "No project"), p.get("company_name", ""),
            round(v["seconds"] / 3600.0, 2), round(v["cost"], 2),
        ])
    _autosize(ws2)

    # Sheet 3: Cost per Employee (monthly total)
    ws3 = wb.create_sheet("Cost per Employee")
    _apply_header(ws3, ["Employee", "Designation", "Range hours", "Range cost (₹)",
                        "Monthly hours", "Monthly cost (₹)"])
    emp_agg = {}
    for s in sessions:
        info = cost_map.get(s.get("user_id"))
        t = tasks.get(s.get("task_id"))
        if not info or not t:
            continue
        if not _pass_filters(s, t):
            continue
        cost, secs = _session_cost_and_seconds(s, info["hourly"], win_start, win_end)
        if secs <= 0:
            continue
        row = emp_agg.setdefault(s["user_id"], {"cost": 0.0, "seconds": 0})
        row["cost"] += cost
        row["seconds"] += secs
    # Also compute monthly totals for each employee
    month_start, month_end, _ = _range_window("month")
    month_sess = await db.timer_sessions.find({
        "$or": [
            {"started_at": {"$gte": month_start.isoformat()}},
            {"ended_at": None},
        ]
    }, {"_id": 0}).to_list(200000)
    monthly = {}
    for s in month_sess:
        info = cost_map.get(s.get("user_id"))
        if not info:
            continue
        cost, secs = _session_cost_and_seconds(s, info["hourly"], month_start, month_end)
        if secs <= 0:
            continue
        row = monthly.setdefault(s["user_id"], {"cost": 0.0, "seconds": 0})
        row["cost"] += cost
        row["seconds"] += secs
    for uid, v in sorted(emp_agg.items(), key=lambda kv: kv[1]["cost"], reverse=True):
        u = users.get(uid, {})
        m = monthly.get(uid, {"cost": 0.0, "seconds": 0})
        ws3.append([
            u.get("first_name", ""), u.get("designation", ""),
            round(v["seconds"] / 3600.0, 2), round(v["cost"], 2),
            round(m["seconds"] / 3600.0, 2), round(m["cost"], 2),
        ])
    _autosize(ws3)

    return _wb_response(wb, "raybotix-costs.xlsx")


@router.get("/productivity.xlsx")
async def export_productivity(user=Depends(require_roles("super_admin", "admin", "manager"))):
    users = await db.users.find({"status": "active"}, {"_id": 0}).to_list(500)
    sessions = await db.timer_sessions.find({}, {"_id": 0}).to_list(50000)
    tasks_map = {}
    for t in await db.tasks.find({}, {"_id": 0, "id": 1, "assignee_id": 1, "status": 1}).to_list(5000):
        tasks_map[t["id"]] = t
    by_user = {}
    for s in sessions:
        by_user.setdefault(s["user_id"], 0)
        by_user[s["user_id"]] += s.get("duration_seconds", 0) or 0

    wb = Workbook()
    ws = wb.active
    ws.title = "Productivity"
    _apply_header(ws, ["Employee", "Designation", "Total hours", "Completed tasks",
                       "Reopened tasks", "Hourly ₹", "Total cost ₹", "Productivity %"])
    for u in users:
        secs = by_user.get(u["id"], 0)
        hours = secs / 3600.0
        wh = u.get("working_hours_per_day", 8) or 8
        wd = u.get("working_days_per_month", 25) or 25
        salary = u.get("monthly_salary", 0) or 0
        mh = wh * wd
        hourly = (salary / mh) if mh else 0
        completed = sum(1 for t in tasks_map.values() if t.get("assignee_id") == u["id"] and t.get("status") == "Completed")
        reopened = sum(1 for t in tasks_map.values() if t.get("assignee_id") == u["id"] and t.get("status") == "Reopened")
        productivity = round(min(100, (secs / (mh * 3600)) * 100), 1) if mh else 0
        ws.append([
            u["first_name"], u.get("designation", ""), round(hours, 2),
            completed, reopened, round(hourly, 2),
            round(hours * hourly, 2), productivity,
        ])
    _autosize(ws)
    return _wb_response(wb, "raybotix-productivity.xlsx")
